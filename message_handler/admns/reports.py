from aiogram import Router, F, Bot
from aiogram.types import Message, ReplyKeyboardRemove, FSInputFile
from aiogram.utils.media_group import MediaGroupBuilder
from aiogram.fsm.context import FSMContext
from aiogram.filters import and_f
from data import sqlPrompts
import keyboards
import words
import filters
import states
import utils
import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from datetime import datetime, timedelta

router = Router()

@router.message(and_f(filters.admins.IsAdmin(), filters.any.textInTuple("Hisobotlar", "Отчеты", "Esabatlar")))
async def my_orders_answer(message: Message, state: FSMContext):
    lan = sqlPrompts.get_user(message.from_user.id)[1]
    text = words.dict[lan]["reports"]["main_text"]
    await message.answer(text, reply_markup=keyboards.reply.reports.admin_menu_builder(lan))
    await state.set_state(states.reports.FilteredReports.type)

@router.message(and_f(states.reports.FilteredReports.type, filters.any.textInTuple("Kunlik", "Ежедневно", "Kúnlik", "Haftalik", "Еженедельно", "Háptelik", "Oylik", "Ежемесячно", "Aylıq", "Yillik", "Ежегодно", "Jıllıq")))
async def my_canceled_orders_answer(message: Message):
    lan = sqlPrompts.get_user(message.from_user.id)[1]
    sana1 = datetime.now()
    sana2 = datetime.now()
    
    if message.text.title() in ("Kunlik", "Ежедневно", "Kúnlik"):
        sana1 -= timedelta(days=1)
    elif message.text.title() in ("Haftalik", "Еженедельно", "Háptelik"):
        sana1 -= timedelta(weeks=1)
    elif message.text.title() in ("Oylik", "Ежемесячно", "Aylıq"):
        sana1 -= timedelta(weeks=4)
    elif message.text.title() in ("Yillik", "Ежегодно", "Jıllıq"):
        sana1 -= timedelta(weeks=56)
    else:
        text = words.dict[lan]["reports"]["alert"]
        await message.answer(text)
        return


    data = sqlPrompts.get_all_report(sana1, sana2)

    if data:    
        wb = openpyxl.Workbook()
        ws = wb.active
            
        ws.append(words.dict[lan]["my_orders"]["my_orders_table_columns"])
        
        for i in "ABCDEFG": ws.column_dimensions[i].width = 20

        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        for i in "ABCDEFG": 
            ws[f'{i}1'].font = Font(color="FFFFFF", bold=True)  # Matn rangini qizil va qalin qilib sozlaymiz
            ws[f'{i}1'].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Orqa fonni yashil rangga sozlaymiz
        
        for i in data:
            product = sqlPrompts.get_productById(i[1])
            if product:
                user = sqlPrompts.get_user(i[5])
                ws.append((
                    str(i[5]),
                    user[0],
                    str(i[0]),
                    product[0], 
                    product[2], 
                    str(i[2]),
                    str(i[2] * product[3]),
                    i[3],
                    "⏳" if i[4] == "wait" else "❌" if i[4] == "rejected" else "✅"
                    ))
            else:
                user = sqlPrompts.get_user(i[5])
                ws.append((
                    str(i[5]),
                    user[0],
                    str(i[0]),
                    "None", 
                    "None", 
                    str(i[2]),
                    "None",
                    i[3],
                    "⏳" if i[4] == "wait" else "❌" if i[4] == "rejected" else "✅"
                ))


        wb.save("reports.xlsx")

        await message.answer_document(FSInputFile("reports.xlsx"))    
    else:
        text = words.dict[lan]["reports"]["not_report"]
        await message.answer(text)

@router.message(and_f(states.reports.FilteredReports.type, filters.any.textInTuple("❌ Rad etilgan", "🕙 Tasdiqlanmagan", "❌ Отклонено", "🕙 Неподтверждено", "❌ Biykarlaw etilgen", "🕙 Tastıyıqlanmagan")))
async def get_reports_filtered_answer(message: Message, state:FSMContext):
    lan = sqlPrompts.get_user(message.from_user.id)[1]
    status = ""
    
    if message.text in ("❌ Rad etilgan", "❌ Отклонено", "❌ Biykarlaw etilgen"):
        status = "rejected"
        
    elif message.text in ("🕙 Tasdiqlanmagan", "🕙 Неподтверждено", "🕙 Tastıyıqlanmagan"):
        status = "wait"
    else:
        text = words.dict[lan]["reports"]["alert"]
        await message.answer(text)
        return

    data = sqlPrompts.get_filtered_all_orders(status)


    if data:    
        wb = openpyxl.Workbook()
        ws = wb.active
            
        ws.append(words.dict[lan]["my_orders"]["my_orders_table_columns"])
        
        for i in "ABCDEFG": ws.column_dimensions[i].width = 20

        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        for i in "ABCDEFG": 
            ws[f'{i}1'].font = Font(color="FFFFFF", bold=True)  # Matn rangini qizil va qalin qilib sozlaymiz
            ws[f'{i}1'].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Orqa fonni yashil rangga sozlaymiz
        
        for i in data:
            product = sqlPrompts.get_productById(i[1])
            if product:
                user = sqlPrompts.get_user(i[5])
                ws.append((
                    str(i[5]),
                    user[0],
                    str(i[0]),
                    product[0], 
                    product[2], 
                    str(i[2]),
                    str(i[2] * product[3]),
                    i[3],
                    "⏳" if i[4] == "wait" else "❌" if i[4] == "rejected" else "✅"
                    ))
            else:
                user = sqlPrompts.get_user(i[5])
                ws.append((
                    str(i[5]),
                    user[0],
                    str(i[0]),
                    "None", 
                    "None", 
                    str(i[2]),
                    "None",
                    i[3],
                    "⏳" if i[4] == "wait" else "❌" if i[4] == "rejected" else "✅"
                ))

        


        wb.save("reports.xlsx")

        await message.answer_document(FSInputFile("reports.xlsx"))    
    else:
        text = words.dict[lan]["reports"]["not_report"]
        await message.answer(text)


@router.message(and_f(states.reports.FilteredReports.type, filters.any.textInTuple("Tovarlar", "Товары", "Tovarlar")))
async def get_reports_filtered_answer(message: Message, state:FSMContext):
    lan = sqlPrompts.get_user(message.from_user.id)[1]
    
    data = sqlPrompts.get_products()

    soni = len(data)
    narx = 0

    for i in data:
        narx += i[2] * i[4]

    if data:    
        wb = openpyxl.Workbook()
        ws = wb.active
            
        ws.append(
            (
                "ID",
                "Имя",
                "Размер",
                "Количество",
                "цена"
            )
        )
        
        for i in "ABCDEFG": ws.column_dimensions[i].width = 20

        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        for i in "ABCDEFG": 
            ws[f'{i}1'].font = Font(color="FFFFFF", bold=True)  # Matn rangini qizil va qalin qilib sozlaymiz
            ws[f'{i}1'].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Orqa fonni yashil rangga sozlaymiz
        
        for i in data:
            ws.append((
                str(i[0]),
                i[1], 
                i[3], 
                str(i[2]),
                str(i[4]) + " USD",
                ))
        


        wb.save("reports.xlsx")

        await message.answer_document(FSInputFile("reports.xlsx"), caption=f"{soni} ta tovar\n💰 {narx} USD")    
    else:
        text = words.dict[lan]["reports"]["not_report"]
        await message.answer(text)
