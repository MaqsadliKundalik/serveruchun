from aiogram import Router, F, Bot
from aiogram.types import Message, ReplyKeyboardRemove, FSInputFile
from aiogram.utils.media_group import MediaGroupBuilder
from aiogram.fsm.context import FSMContext
from aiogram.filters import and_f
from data import sqlPrompts
import keyboards
import words
import filters
import states
import utils
import openpyxl
from openpyxl.styles import Font, Color, PatternFill, Border, Side
from datetime import datetime, timedelta

router = Router()

@router.message(filters.any.textInTuple("📬 Mening buyurtmalarim", "📬 Мои заказы", "📬 Meniń buyirtmalarim"))
async def my_orders_answer(message: Message, state: FSMContext):
    lan = sqlPrompts.get_user(message.from_user.id)[1]
    text = words.dict[lan]["my_orders"]["main_text"]
    await message.answer(text, reply_markup=keyboards.reply.my_orders.orders_type_menu_builder(lan))
    await state.set_state(states.orders.GetFilteredOrders.type)

@router.message(and_f(states.orders.GetFilteredOrders.type, filters.any.textInTuple("🆕 Yangi", "🆕 Новое", "🆕 Jańa")))
async def get_filtered_orders(message: Message):
    lan = sqlPrompts.get_user(message.from_user.id)[1]
    sana1 = datetime.now() - timedelta(weeks=1)
    sana2 = datetime.now()
    
    data = sqlPrompts.get_filtered_orders(message.from_user.id, status="new", date1=sana1, date2=sana2)
    if data:    
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(words.dict[lan]["my_orders"]["my_orders_table_columns"])
        
        for i in "ABCDEFG": ws.column_dimensions[i].width = 20

        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        for i in "ABCDEFG": 
            ws[f'{i}1'].font = Font(color="FFFFFF", bold=True)  # Matn rangini qizil va qalin qilib sozlaymiz
            ws[f'{i}1'].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Orqa fonni yashil rangga sozlaymiz
        product = sqlPrompts.get_productById(data[1])
        if product:
            ws.append((
                str(data[0]),
                product[0], 
                product[2], 
                str(data[2]),
                str(data[2] * product[3]),
                data[3],
                "⏳" if data[4] == "wait" else "❌" if data[4] == "rejected" else "✅"
                ))
        else:
            ws.append((
                str(data[0]),
                "None", 
                "None", 
                str(data[2]),
                "None",
                data[3],
                "⏳" if data[4] == "wait" else "❌" if data[4] == "rejected" else "✅"
                ))
        
        wb.save("new_orders.xlsx")

        await message.answer_document(FSInputFile("new_orders.xlsx"))    
    else:
        text = words.dict[lan]["my_orders"]["not_new_order"]
        await message.answer(text)

@router.message(and_f(states.orders.GetFilteredOrders.type, filters.any.textInTuple("❌ Rad etilgan", "❌ Отклонено", "❌ Biykarlaw etilgen")))
async def my_canceled_orders_answer(message: Message):
    lan = sqlPrompts.get_user(message.from_user.id)[1]

    data = sqlPrompts.get_filtered_orders(message.from_user.id, status="rejected")
    if data:    
        wb = openpyxl.Workbook()
        ws = wb.active
            
        ws.append(words.dict[lan]["my_orders"]["my_orders_table_columns"])
        
        for i in "ABCDEFG": ws.column_dimensions[i].width = 20

        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        for i in "ABCDEFG": 
            ws[f'{i}1'].font = Font(color="FFFFFF", bold=True)  # Matn rangini qizil va qalin qilib sozlaymiz
            ws[f'{i}1'].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Orqa fonni yashil rangga sozlaymiz
        
        for i in data:
            product = sqlPrompts.get_productById(i[1])
            if product:
                ws.append((
                    str(i[0]),
                    product[0], 
                    product[2], 
                    str(i[2]),
                    str(i[2] * product[3]),
                    i[3],
                    "⏳" if i[4] == "wait" else "❌" if i[4] == "rejected" else "✅"
                    ))
            else:
                ws.append((
                    str(i[0]),
                    "None", 
                    "None", 
                    str(i[2]),
                    "None",
                    i[3],
                    "⏳" if i[4] == "wait" else "❌" if i[4] == "rejected" else "✅"
                    ))


        wb.save("canceled_orders.xlsx")

        await message.answer_document(FSInputFile("canceled_orders.xlsx"))    
    else:
        text = words.dict[lan]["my_orders"]["not_canceled_orders"]
        await message.answer(text)

@router.message(and_f(states.orders.GetFilteredOrders.type, filters.any.textInTuple("🕙 Tasdiqlanmagan", "🕙 Неподтверждено", "🕙 Tastıyıqlanmagan")))
async def my_waited_orders_answer(message: Message):
    lan = sqlPrompts.get_user(message.from_user.id)[1]

    data = sqlPrompts.get_filtered_orders(message.from_user.id, status="wait")
    if data:    
        wb = openpyxl.Workbook()
        ws = wb.active
            
        ws.append(words.dict[lan]["my_orders"]["my_orders_table_columns"])
        
        for i in "ABCDEFG": ws.column_dimensions[i].width = 20

        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        for i in "ABCDEFG": 
            ws[f'{i}1'].font = Font(color="FFFFFF", bold=True)  # Matn rangini qizil va qalin qilib sozlaymiz
            ws[f'{i}1'].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Orqa fonni yashil rangga sozlaymiz
        
        for i in data:
            product = sqlPrompts.get_productById(i[1])
            if product:
                ws.append((
                    str(i[0]),
                    product[0], 
                    product[2], 
                    str(i[2]),
                    str(i[2] * product[3]),
                    i[3],
                    "⏳" if i[4] == "wait" else "❌" if i[4] == "rejected" else "✅"
                    ))
            else:
                ws.append((
                    str(i[0]),
                    "None", 
                    "None", 
                    str(i[2]),
                    "None",
                    i[3],
                    "⏳" if i[4] == "wait" else "❌" if i[4] == "rejected" else "✅"
                    ))



        wb.save("pending_orders.xlsx")

        await message.answer_document(FSInputFile("pending_orders.xlsx"))    
    else:
        text = words.dict[lan]["my_orders"]["not_pending_orders"]
        await message.answer(text)
