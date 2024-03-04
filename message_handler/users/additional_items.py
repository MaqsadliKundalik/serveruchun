from aiogram import Router, F
from aiogram.types import Message, FSInputFile
from requests import get
import words
from data import sqlPrompts
from keyboards import reply
from filters import any
import openpyxl 
from openpyxl.styles import Font, Color, PatternFill, Border, Side

router = Router()

@router.message(any.textInTuple("📂 Qosimsha narsalar", "📂 Аксессуары", "📂 Qosimsha zatlar"))
async def additional_items_answer(message: Message):
    lan = sqlPrompts.get_user(message.from_user.id)[1]
    text = words.dict[lan]["additional_items"]["main_text"]

    if lan == "UZ": await message.answer(text, reply_markup=reply.additional_items_menu.additional_items_uz)
    if lan == "RU": await message.answer(text, reply_markup=reply.additional_items_menu.additional_items_ru)
    if lan == "QA": await message.answer(text, reply_markup=reply.additional_items_menu.additional_items_qa)

@router.message(any.textInTuple("💲USD stul", "💲Курс доллара США", "💲USD kursi"))
async def usd_kursov_answer(message: Message):
    getvalyut = get("https://cbu.uz/uz/arkhiv-kursov-valyut/json/")

    Rate = getvalyut.json()[0]["Rate"]
    Date = getvalyut.json()[0]["Date"]

    lan = sqlPrompts.get_user(message.from_user.id)[1]
    text = words.dict[lan]["additional_items"]["dollar_kursi"](Rate, Date)

    await message.answer(text)



@router.message(any.textInTuple("⚖️ Balans", "⚖️ Баланс", "⚖️ Balans"))
async def usd_kursov_answer(message: Message):
    user = sqlPrompts.get_user(message.from_user.id)
    lan = sqlPrompts.get_user(message.from_user.id)[1]
    text = words.dict[lan]["additional_items"]["balans"](user[3])

    await message.answer(text)

@router.message(any.textInTuple("🧾 narxlar ro'yhati", "🧾 прайс-лист", "🧾 Bahalar kestesi"))
async def price_list_answer(message: Message):
    lan = sqlPrompts.get_user(message.from_user.id)[1]
    
    data = sqlPrompts.get_products()

    if data:    
        wb = openpyxl.Workbook()
        ws = wb.active
            
        ws.append(
            (
                "Имя",
                "Размер",
                "цена"
            )
        )
        
        for i in "ABC": ws.column_dimensions[i].width = 20

        border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

        for i in "ABC": 
            ws[f'{i}1'].font = Font(color="FFFFFF", bold=True)  # Matn rangini qizil va qalin qilib sozlaymiz
            ws[f'{i}1'].fill = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")  # Orqa fonni yashil rangga sozlaymiz
        
        for i in data:
            ws.append((
                i[1], 
                i[3], 
                str(i[4]) + " USD",
                ))
        


        wb.save("price_list.xlsx")

        await message.answer_document(FSInputFile("price_list.xlsx"))    
    else:
        await message.answer("🤔")
